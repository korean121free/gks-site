# -*- coding: utf-8 -*-
"""
1365 자원봉사활동 인증신청서 만들기 — 성남시 공식 엑셀 서식

Worker에서 그 달 신청 자료를 받아, 성남시 공식 서식을 '템플릿'으로 열어
값만 채웁니다. 그래서 표지 서식과 단체 직인 이미지가 그대로 보존됩니다.

만들어지는 것: 제출용 ZIP 하나
    글로벌한국어나눔26-07/
      글로벌한국어나눔(26-7월).xlsx     ← 표지 + 봉사참여명단
      증빙사진/김선영2026.07.09_Linh.jpg
      한글보고서/김윤희-7월봉사신청서.hwp   (한글로 내신 분만)

서식의 '작성 유의사항' 규칙을 코드로 지킵니다.
    · 생년월일 / 활동일자 : YYYY-MM-DD
    · 시작·마침 시간      : HH:MM, 24시간제, 10분 단위
    · 공란 없이 작성 (반복되는 내용도 매 행에 다시 씀)
    · 개인정보 활용동의 체크 필수
    · 활동인원 = 연인원(누적 연번) = 명단 행 수

쓰는 법 (1365만들기.bat 이 대신 실행해 줍니다)
    python make_1365.py 2026-07
"""
import sys, os, json, shutil, zipfile, datetime, urllib.request, urllib.parse
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
TEMPLATE = HERE / '1365서식.xlsx'        # 성남시 공식 빈 서식 (직인 포함)
OUTDIR = HERE / '제출'

ORG = '글로벌한국어나눔'
CHECK = '☑'
FIRST_ROW = 5                            # 봉사참여명단 첫 데이터 행

API = os.environ.get('CLASS_API', '').rstrip('/')
KEY = os.environ.get('ADMIN_KEY', '')


# ---------------------------------------------------------------- 자료 받기

def fetch(month: str) -> dict:
    if not API or not KEY:
        sys.exit('CLASS_API / ADMIN_KEY 환경변수가 없습니다. 1365만들기.bat 로 실행해 주세요.')
    url = f'{API}/form1365?key={urllib.parse.quote(KEY)}&month={month}'
    with urllib.request.urlopen(url, timeout=30) as r:
        d = json.load(r)
    if d.get('error'):
        sys.exit(f'자료를 받지 못했습니다: {d["error"]}')
    return d


def download(rkey: str, dest: Path):
    url = f'{API}/photo/get?k={urllib.parse.quote(rkey)}&dl=1'
    dest.parent.mkdir(parents=True, exist_ok=True)
    with urllib.request.urlopen(url, timeout=60) as r:
        # 서버가 알려 주는 파일 이름을 그대로 씁니다
        cd = r.headers.get('Content-Disposition', '')
        name = None
        if "filename*=UTF-8''" in cd:
            name = urllib.parse.unquote(cd.split("filename*=UTF-8''")[1].strip('";'))
        out = dest.parent / (name or dest.name)
        out.write_bytes(r.read())
    return out


# ---------------------------------------------------------------- 엑셀 채우기

def r10(hhmm: str) -> datetime.time:
    """10분 단위로 맞춥니다 (서식 요구사항)."""
    t = datetime.datetime.strptime(hhmm, '%H:%M').time()
    return datetime.time(t.hour, (t.minute // 10) * 10)


def d(s: str) -> datetime.datetime:
    return datetime.datetime.strptime(s[:10], '%Y-%m-%d')


def build_xlsx(data: dict, out: Path) -> tuple:
    """명단을 채운 엑셀을 만들고 (경로, 채운행수, 넘친행수) 를 돌려줍니다."""
    if not TEMPLATE.exists():
        sys.exit(f'서식 파일이 없습니다: {TEMPLATE}\n'
                 f'성남시 홈페이지(서식자료실)에서 받은 엑셀에 직인을 넣어 이 이름으로 두세요.')

    # 활동 1건 = 1행. 성명 → 날짜 순.
    rows = []
    for T in data['teachers']:
        for r in T['rows']:
            rows.append({
                'name': T['name'], 'birth': T['birth'],
                'pid': T['portal_id'] or T['phone'],
                'day': r['day'], 'start': r['start'], 'end': r['end'],
                'place': r['place'], 'content': r['content'],
            })
    rows.sort(key=lambda x: (x['name'], x['day'], x['start']))

    shutil.copy(TEMPLATE, out)
    wb = openpyxl.load_workbook(out)

    month = data['month']
    y, m = int(month[:4]), int(month[5:7])
    last = (datetime.date(y + (m == 12), (m % 12) + 1, 1) - datetime.timedelta(days=1)).day

    # ---- 표지 ----
    cv = wb['표지']
    cv['E6'] = f'{y}     .   {m:02d}  . 01   .    ~   {y}  .   {m:02d} .   {last}'
    cv['E8'] = f'            {len(rows)}  명'          # 연인원 = 행 수
    cv['E10'] = '한국어 튜터링'
    t = datetime.date.today()
    w = data['writer']
    cv['H17'] = f'        {t.year}    년  {t.month} 월   {t.day}   일'
    cv['H18'] = f"(직책) {w['title']}         (성명) {w['name']}"
    cv['H19'] = w['phone']
    cv['A22'] = f'단체(기관)명 :     {ORG}'

    # ---- 봉사참여명단 ----
    ws = wb['봉사참여명단']
    ws['A1'] = f'■ 단체(기관)명 : {ORG}'
    capacity = ws.max_row - FIRST_ROW + 1
    over = max(0, len(rows) - capacity)

    for i in range(capacity):
        r = FIRST_ROW + i
        if i < len(rows):
            x = rows[i]
            ws.cell(r, 1).value = i + 1
            ws.cell(r, 2).value = x['name']                       # 매 행 반복 (공란 금지)
            ws.cell(r, 3).value = d(x['birth']) if x['birth'] else None
            ws.cell(r, 4).value = x['pid']
            ws.cell(r, 5).value = d(x['day'])
            ws.cell(r, 6).value = r10(x['start'])
            ws.cell(r, 7).value = r10(x['end'])
            ws.cell(r, 8).value = x['place']
            ws.cell(r, 9).value = x['content']
            ws.cell(r, 10).value = CHECK
        else:
            for c in range(2, 10):
                ws.cell(r, c).value = None
            ws.cell(r, 10).value = '□'
        ws.cell(r, 3).number_format = 'yyyy-mm-dd'
        ws.cell(r, 5).number_format = 'yyyy-mm-dd'
        ws.cell(r, 6).number_format = 'hh:mm'
        ws.cell(r, 7).number_format = 'hh:mm'

    wb.save(out)
    return out, min(len(rows), capacity), over


def audit(path: Path) -> list:
    """만든 파일을 다시 열어 규칙 위반을 스스로 점검합니다."""
    ws = openpyxl.load_workbook(path)['봉사참여명단']
    bad = []
    for r in range(FIRST_ROW, ws.max_row + 1):
        name = ws.cell(r, 2).value
        if not name:
            continue
        if not ws.cell(r, 3).value:
            bad.append(f'{r}행 {name}: 생년월일 없음')
        if not ws.cell(r, 4).value:
            bad.append(f'{r}행 {name}: 1365포털ID/연락처 없음 — 인증 처리 불가')
        for c, label in ((5, '활동일자'), (6, '시작시간'), (7, '마침시간'), (8, '장소'), (9, '활동내용')):
            if not ws.cell(r, c).value:
                bad.append(f'{r}행 {name}: {label} 공란')
        for c, label in ((6, '시작시간'), (7, '마침시간')):
            v = ws.cell(r, c).value
            if isinstance(v, datetime.time) and v.minute % 10:
                bad.append(f'{r}행 {name}: {label} 10분 단위 아님 ({v})')
        if ws.cell(r, 10).value != CHECK:
            bad.append(f'{r}행 {name}: 개인정보 활용동의 체크 없음')
    return bad


# ---------------------------------------------------------------- 실행

def main():
    month = sys.argv[1] if len(sys.argv) > 1 else datetime.date.today().strftime('%Y-%m')
    print(f'\n{month} 1365 인증신청서를 만듭니다.\n')

    data = fetch(month)
    s = data['summary']
    print(f"  신청한 선생님 {s['teachers']}명 / 그 달 수업한 선생님 {s['taught']}명")
    if s['not_requested']:
        print(f"  · 신청을 안 누른 선생님 {s['not_requested']}명 (선택이므로 제외됩니다)")
    if not data['teachers']:
        sys.exit('  신청한 선생님이 없습니다. 선생님 방에서 [1365 신청]을 누르시면 여기에 나옵니다.')

    folder = OUTDIR / f'{ORG}{month[2:4]}-{int(month[5:7])}'
    if folder.exists():
        shutil.rmtree(folder)
    folder.mkdir(parents=True)

    xlsx, filled, over = build_xlsx(data, folder / f'{ORG}({month[2:4]}-{int(month[5:7])}월).xlsx')
    print(f'  명단 {filled}건 채움 (연인원 {filled}명)')
    if over:
        print(f'  ⚠ {over}건이 서식 행 수를 넘었습니다 — 센터에 문의가 필요합니다 (031-757-6226)')

    # 증빙 사진 · 한글 보고서
    photos = reports = 0
    for T in data['teachers']:
        if T.get('photo_key'):
            download(T['photo_key'], folder / '증빙사진' / 'x.jpg'); photos += 1
        for rp in T.get('reports', []):
            download(rp['rkey'], folder / '한글보고서' / 'x.hwp'); reports += 1
    print(f'  증빙 사진 {photos}장 · 한글 보고서 {reports}개')

    # 점검
    bad = audit(xlsx)
    print('\n  자동 점검:', '이상 없음' if not bad else f'{len(bad)}건 — 보내기 전에 확인하세요')
    for x in bad[:15]:
        print('   -', x)

    # ZIP
    zp = OUTDIR / f'{folder.name}.zip'
    with zipfile.ZipFile(zp, 'w', zipfile.ZIP_DEFLATED) as z:
        for p in sorted(folder.rglob('*')):
            if p.is_file():
                z.write(p, p.relative_to(OUTDIR))
    print(f'\n  만들었습니다: {zp}')
    print(f'  받는 곳: snvol@hanmail.net  (fax 031-757-6229)')
    print('  ※ 열어서 점검하신 뒤 직접 보내 주세요. 메일 발송은 하지 않습니다.\n')

    os.startfile(OUTDIR)   # 폴더를 열어 드립니다


if __name__ == '__main__':
    main()
